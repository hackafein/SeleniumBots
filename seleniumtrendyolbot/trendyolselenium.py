import os, time, random
from selenium import webdriver
import xlsxwriter
from selenium.webdriver.chrome.options import Options
import timeit
import re
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook
from openpyxl import load_workbook


class Ui_MainWindow(object):
    def __init__(self):
        self.path = f'{os.getcwd()}/chromedriver'
        self.chrome_options = Options()
        mobile_emulation = {
        "deviceMetrics": {"width": 360, "height": 640, "pixelRatio": 3.0},
        "userAgent": "Mozilla/5.0 (Linux; Android 4.2.1; en-us; Nexus 5 Build/JOP40D) AppleWebKit/535.19 (KHTML, like Gecko) Chrome/18.0.1025.166 Mobile Safari/535.19"}
        #self.chrome_options.add_argument("--headless")
        self.chrome_options.add_experimental_option("mobileEmulation", mobile_emulation)
        self.driver = webdriver.Chrome(self.path, chrome_options=self.chrome_options)
        self.verilist=[]
        self.userlist=["a.n.k.a.r.a.ha.y.mana.li2020@gmail.com","a.nk.arahay.man.ali2020@gmail.com"]
        self.passlist=["Qwert678","Qwert678"]
        self.step=0
        self.siparis_step=0
        self.siparis_verilist=[]
        self.bigstep=0
        self.bigstep_siparis=0
        self.chngpss_step=0
        self.username=""
        self.passw=""
        self.cond=True

    def setupUi(self, MainWindow):

        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 600)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setGeometry(QtCore.QRect(0, 0, 791, 551))
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.pushButton = QtWidgets.QPushButton(self.tab)
        self.pushButton.setGeometry(QtCore.QRect(600, 450, 113, 32))
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.couponcrawl)
        self.couponprogressBar = QtWidgets.QProgressBar(self.tab)
        self.couponprogressBar.setGeometry(QtCore.QRect(200, 450, 391, 21))
        self.couponprogressBar.setProperty("value", 0)
        self.couponprogressBar.setObjectName("couponprogressBar")
        self.usernametextedit = QtWidgets.QPlainTextEdit(self.tab)
        self.usernametextedit.setGeometry(QtCore.QRect(50, 70, 311, 261))
        self.usernametextedit.setObjectName("usernametextedit")
        self.passtextedit = QtWidgets.QPlainTextEdit(self.tab)
        self.passtextedit.setGeometry(QtCore.QRect(440, 70, 321, 261))
        self.passtextedit.setObjectName("passtextedit")
        self.label = QtWidgets.QLabel(self.tab)
        self.label.setGeometry(QtCore.QRect(150, 40, 101, 21))
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.tab)
        self.label_2.setGeometry(QtCore.QRect(580, 40, 101, 21))
        self.label_2.setObjectName("label_2")
        self.lineEdit = QtWidgets.QLineEdit(self.tab)
        self.lineEdit.setGeometry(QtCore.QRect(340, 380, 113, 21))
        self.lineEdit.setObjectName("lineEdit")
        self.label_3 = QtWidgets.QLabel(self.tab)
        self.label_3.setGeometry(QtCore.QRect(230, 380, 111, 20))
        self.label_3.setObjectName("label_3")
        self.teksifre = QtWidgets.QCheckBox(self.tab)
        self.teksifre.setGeometry(QtCore.QRect(600, 340, 161, 21))
        self.teksifre.setObjectName("teksifre")
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tabWidget.addTab(self.tab, "")
        self.siparislericek = QtWidgets.QWidget()
        self.siparislericek.setObjectName("siparislericek")
        self.lineEdit_siparis = QtWidgets.QLineEdit(self.siparislericek)
        self.lineEdit_siparis.setGeometry(QtCore.QRect(320, 370, 113, 21))
        self.lineEdit_siparis.setObjectName("lineEdit_siparis")
        self.exellabelsiparis = QtWidgets.QLabel(self.siparislericek)
        self.exellabelsiparis.setGeometry(QtCore.QRect(210, 370, 111, 20))
        self.exellabelsiparis.setObjectName("exellabelsiparis")
        self.sipariskullaniciadilabel = QtWidgets.QLabel(self.siparislericek)
        self.sipariskullaniciadilabel.setGeometry(QtCore.QRect(130, 30, 101, 21))
        self.sipariskullaniciadilabel.setObjectName("sipariskullaniciadilabel")
        self.siparissifrelabel = QtWidgets.QLabel(self.siparislericek)
        self.siparissifrelabel.setGeometry(QtCore.QRect(560, 30, 101, 21))
        self.siparissifrelabel.setObjectName("siparissifrelabel")
        self.pushButton_siparis = QtWidgets.QPushButton(self.siparislericek)
        self.pushButton_siparis.setGeometry(QtCore.QRect(580, 440, 113, 32))
        self.pushButton_siparis.setObjectName("pushButton_siparis")
        self.pushButton_siparis.clicked.connect(self.getorderspyqt)
        self.passtextedit_siparis = QtWidgets.QPlainTextEdit(self.siparislericek)
        self.passtextedit_siparis.setGeometry(QtCore.QRect(420, 60, 321, 261))
        self.passtextedit_siparis.setObjectName("passtextedit_siparis")
        self.couponprogressBar_siparis = QtWidgets.QProgressBar(self.siparislericek)
        self.couponprogressBar_siparis.setGeometry(QtCore.QRect(180, 440, 391, 21))
        self.couponprogressBar_siparis.setProperty("value", 0)
        self.couponprogressBar_siparis.setObjectName("couponprogressBar_siparis")
        self.usernametextedit_siparis = QtWidgets.QPlainTextEdit(self.siparislericek)
        self.usernametextedit_siparis.setGeometry(QtCore.QRect(30, 60, 311, 261))
        self.usernametextedit_siparis.setObjectName("usernametextedit_siparis")
        self.teksifre_siparis = QtWidgets.QCheckBox(self.siparislericek)
        self.teksifre_siparis.setGeometry(QtCore.QRect(580, 330, 161, 21))
        self.teksifre_siparis.setObjectName("teksifre_siparis")
        self.tabWidget.addTab(self.siparislericek, "")

        self.tab_3 = QtWidgets.QWidget()
        self.tab_3.setObjectName("tab_3")
        self.label_4 = QtWidgets.QLabel(self.tab_3)
        self.label_4.setGeometry(QtCore.QRect(100, 20, 101, 21))
        self.label_4.setObjectName("label_4")
        self.usernametextedit_chngpss = QtWidgets.QPlainTextEdit(self.tab_3)
        self.usernametextedit_chngpss.setGeometry(QtCore.QRect(40, 50, 251, 261))
        self.usernametextedit_chngpss.setObjectName("usernametextedit_chngpss")
        self.label_5 = QtWidgets.QLabel(self.tab_3)
        self.label_5.setGeometry(QtCore.QRect(360, 20, 101, 21))
        self.label_5.setObjectName("label_5")
        self.passtextedit_chngoldpss = QtWidgets.QPlainTextEdit(self.tab_3)
        self.passtextedit_chngoldpss.setGeometry(QtCore.QRect(310, 50, 211, 261))
        self.passtextedit_chngoldpss.setObjectName("passtextedit_chngoldpss")
        self.teksifre_chngoldpss = QtWidgets.QCheckBox(self.tab_3)
        self.teksifre_chngoldpss.setGeometry(QtCore.QRect(340, 320, 161, 21))
        self.teksifre_chngoldpss.setObjectName("teksifre_chngoldpss")
        self.pushButton_chngpss = QtWidgets.QPushButton(self.tab_3)
        self.pushButton_chngpss.setGeometry(QtCore.QRect(550, 450, 141, 41))
        self.pushButton_chngpss.setObjectName("pushButton_chngpss")
        self.pushButton_chngpss.clicked.connect(self.pyqtchangepass)
        self.couponprogressBar_chngpss = QtWidgets.QProgressBar(self.tab_3)
        self.couponprogressBar_chngpss.setGeometry(QtCore.QRect(140, 460, 391, 21))
        self.couponprogressBar_chngpss.setProperty("value", 0)
        self.couponprogressBar_chngpss.setObjectName("couponprogressBar_chngpss")
        self.teksifre_chngnewpss = QtWidgets.QCheckBox(self.tab_3)
        self.teksifre_chngnewpss.setGeometry(QtCore.QRect(580, 320, 161, 21))
        self.teksifre_chngnewpss.setObjectName("teksifre_chngnewpss")
        self.passtextedit_chngnewpss = QtWidgets.QPlainTextEdit(self.tab_3)
        self.passtextedit_chngnewpss.setGeometry(QtCore.QRect(560, 50, 211, 261))
        self.passtextedit_chngnewpss.setObjectName("passtextedit_chngnewpss")
        self.label_6 = QtWidgets.QLabel(self.tab_3)
        self.label_6.setGeometry(QtCore.QRect(610, 20, 101, 21))
        self.label_6.setObjectName("label_6")
        self.tabWidget.addTab(self.tab_3, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.label_7 = QtWidgets.QLabel(self.tab_2)
        self.label_7.setGeometry(QtCore.QRect(520, 10, 101, 21))
        self.label_7.setObjectName("label_7")
        self.usernametextedit_adres = QtWidgets.QPlainTextEdit(self.tab_2)
        self.usernametextedit_adres.setGeometry(QtCore.QRect(20, 30, 331, 261))
        self.usernametextedit_adres.setObjectName("usernametextedit_adres")
        self.label_8 = QtWidgets.QLabel(self.tab_2)
        self.label_8.setGeometry(QtCore.QRect(80, 0, 101, 21))
        self.label_8.setObjectName("label_8")
        self.passtextedit_adres = QtWidgets.QPlainTextEdit(self.tab_2)
        self.passtextedit_adres.setGeometry(QtCore.QRect(440, 30, 311, 261))
        self.passtextedit_adres.setObjectName("passtextedit_adres")
        self.teksifre_adres = QtWidgets.QCheckBox(self.tab_2)
        self.teksifre_adres.setGeometry(QtCore.QRect(450, 290, 161, 21))
        self.teksifre_adres.setObjectName("teksifre_adres")
        self.lineEdit_adresbasligi = QtWidgets.QLineEdit(self.tab_2)
        self.lineEdit_adresbasligi.setGeometry(QtCore.QRect(140, 340, 151, 21))
        self.lineEdit_adresbasligi.setObjectName("lineEdit_adresbasligi")
        self.lineEdit_adresad = QtWidgets.QLineEdit(self.tab_2)
        self.lineEdit_adresad.setGeometry(QtCore.QRect(80, 370, 113, 21))
        self.lineEdit_adresad.setObjectName("lineEdit_adresad")
        self.lineEdit_adressoyad = QtWidgets.QLineEdit(self.tab_2)
        self.lineEdit_adressoyad.setGeometry(QtCore.QRect(280, 370, 113, 21))
        self.lineEdit_adressoyad.setObjectName("lineEdit_adressoyad")
        self.label_9 = QtWidgets.QLabel(self.tab_2)
        self.label_9.setGeometry(QtCore.QRect(40, 340, 91, 16))
        self.label_9.setObjectName("label_9")
        self.label_10 = QtWidgets.QLabel(self.tab_2)
        self.label_10.setGeometry(QtCore.QRect(40, 370, 91, 16))
        self.label_10.setObjectName("label_10")
        self.label_11 = QtWidgets.QLabel(self.tab_2)
        self.label_11.setGeometry(QtCore.QRect(220, 370, 61, 16))
        self.label_11.setObjectName("label_11")
        self.label_12 = QtWidgets.QLabel(self.tab_2)
        self.label_12.setGeometry(QtCore.QRect(40, 410, 101, 16))
        self.label_12.setObjectName("label_12")
        self.lineEdit_adrestel = QtWidgets.QLineEdit(self.tab_2)
        self.lineEdit_adrestel.setGeometry(QtCore.QRect(140, 410, 151, 21))
        self.lineEdit_adrestel.setObjectName("lineEdit_adrestel")
        self.textEdit_adres = QtWidgets.QTextEdit(self.tab_2)
        self.textEdit_adres.setGeometry(QtCore.QRect(500, 330, 261, 101))
        self.textEdit_adres.setObjectName("textEdit_adres")
        self.label_13 = QtWidgets.QLabel(self.tab_2)
        self.label_13.setGeometry(QtCore.QRect(590, 310, 61, 16))
        self.label_13.setObjectName("label_13")
        self.pushButton_adresekle = QtWidgets.QPushButton(self.tab_2)
        self.pushButton_adresekle.setGeometry(QtCore.QRect(640, 490, 113, 32))
        self.pushButton_adresekle.setObjectName("pushButton_adresekle")
        self.couponprogressBar_adres = QtWidgets.QProgressBar(self.tab_2)
        self.couponprogressBar_adres.setGeometry(QtCore.QRect(370, 490, 261, 21))
        self.couponprogressBar_adres.setProperty("value", 0)
        self.couponprogressBar_adres.setObjectName("couponprogressBar_adres")
        self.label_14 = QtWidgets.QLabel(self.tab_2)
        self.label_14.setGeometry(QtCore.QRect(40, 450, 101, 16))
        self.label_14.setObjectName("label_14")
        self.lineEdit_adresil = QtWidgets.QLineEdit(self.tab_2)
        self.lineEdit_adresil.setGeometry(QtCore.QRect(70, 450, 111, 21))
        self.lineEdit_adresil.setObjectName("lineEdit_adresil")
        self.lineEdit_adresilce = QtWidgets.QLineEdit(self.tab_2)
        self.lineEdit_adresilce.setGeometry(QtCore.QRect(240, 450, 111, 21))
        self.lineEdit_adresilce.setText("")
        self.lineEdit_adresilce.setObjectName("lineEdit_adresilce")
        self.label_15 = QtWidgets.QLabel(self.tab_2)
        self.label_15.setGeometry(QtCore.QRect(210, 450, 101, 16))
        self.label_15.setObjectName("label_15")
        self.lineEdit_adresmahalle = QtWidgets.QLineEdit(self.tab_2)
        self.lineEdit_adresmahalle.setGeometry(QtCore.QRect(92, 490, 131, 21))
        self.lineEdit_adresmahalle.setObjectName("lineEdit_adresmahalle")
        self.label_16 = QtWidgets.QLabel(self.tab_2)
        self.label_16.setGeometry(QtCore.QRect(30, 490, 101, 16))
        self.label_16.setObjectName("label_16")
        self.tabWidget.addTab(self.tab_2, "")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 24))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        
        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Trendyol Crawler"))
        self.pushButton.setText(_translate("MainWindow", "Kuponları Çek"))
        self.label.setText(_translate("MainWindow", "Kullanıcı adları"))
        self.label_2.setText(_translate("MainWindow", "Şifreler"))
        self.label_3.setText(_translate("MainWindow", "Excel dosya adı ="))
        self.teksifre.setText(_translate("MainWindow", "Hepsi için tek şifre"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("MainWindow", "Kuponları Çek"))
        self.exellabelsiparis.setText(_translate("MainWindow", "Excel dosya adı ="))
        self.sipariskullaniciadilabel.setText(_translate("MainWindow", "Kullanıcı adları"))
        self.siparissifrelabel.setText(_translate("MainWindow", "Şifreler"))
        self.pushButton_siparis.setText(_translate("MainWindow", "Siparişleri Çek"))
        self.teksifre_siparis.setText(_translate("MainWindow", "Hepsi için tek şifre"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.siparislericek), _translate("MainWindow", "Siparişleri Çek"))
        self.label_4.setText(_translate("MainWindow", "Kullanıcı adları"))
        self.label_5.setText(_translate("MainWindow", "Eski Şifreler"))
        self.teksifre_chngoldpss.setText(_translate("MainWindow", "Hepsi için tek şifre"))
        self.pushButton_chngpss.setText(_translate("MainWindow", "Şifreleri Değiştir"))
        self.teksifre_chngnewpss.setText(_translate("MainWindow", "Hepsi için tek şifre koy"))
        self.label_6.setText(_translate("MainWindow", "Yeni Şifreler"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_3), _translate("MainWindow", "Şifre Değiştir"))
        self.label_7.setText(_translate("MainWindow", "Şifreler"))
        self.label_8.setText(_translate("MainWindow", "Kullanıcı adları"))
        self.teksifre_adres.setText(_translate("MainWindow", "Hepsi için tek şifre"))
        self.label_9.setText(_translate("MainWindow", "Adres Başlığı="))
        self.label_10.setText(_translate("MainWindow", "Ad ="))
        self.label_11.setText(_translate("MainWindow", "Soyad ="))
        self.label_12.setText(_translate("MainWindow", "Cep Telefonu ="))
        self.label_13.setText(_translate("MainWindow", "Adres="))
        self.pushButton_adresekle.setText(_translate("MainWindow", "Adres Ekle"))
        self.label_14.setText(_translate("MainWindow", "İl ="))
        self.label_15.setText(_translate("MainWindow", "İlçe="))
        self.label_16.setText(_translate("MainWindow", "Mahalle ="))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("MainWindow", "Adres Ekle"))
    def couponcrawl(self):
        usertext = self.usernametextedit.toPlainText()
        passtext = self.passtextedit.toPlainText()
        self.userlist=usertext.split('\n')
        self.passlist=passtext.split('\n')
        self.couponcrawler(self.userlist,self.passlist)
        self.couponprogressBar.setMaximum(len(self.userlist))
    def getorderspyqt(self):
        usertext = self.usernametextedit_siparis.toPlainText()
        passtext = self.passtextedit_siparis.toPlainText()
        self.userlist=usertext.split('\n')
        self.passlist=passtext.split('\n')
        self.getallorders(self.userlist,self.passlist)
        self.couponprogressBar_siparis.setMaximum(len(self.userlist))
    def login(self):
        try:
            self.driver.get(str("https://m.trendyol.com/giris?cb=%2F"))

            userelement= WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div[2]/form/div[1]/input')))

            userelement.send_keys(self.username)
            
            passelement=self.driver.find_element_by_xpath('/html/body/main/div/div[2]/form/div[2]/div/div/input')
            passelement.send_keys(self.passw)
            button= WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div[2]/form/button')))
            button.click()
        
            time.sleep(4)
            try:
                self.driver.find_element_by_xpath("/html/body/main/div/div[2]/span[2]")
                self.verilist.append(["GIRIS BASARISIZ","GIRIS BASARISIZ","GIRIS BASARISIZ","GIRIS BASARISIZ","GIRIS BASARISIZ"])
                #self.kuponcrawler()
                return False
            except:
                return True
        except:
            self.login()
            
            
    def kuponcrawler(self):
        self.driver.get(str("https://m.trendyol.com/hesabim/indirim-kuponlari"))
        
        try:
            nocupon = WebDriverWait(self.driver, 3).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div/div/div[2]/div/p')))
            self.verilist.append(["KUPON YOK","KUPON YOK","KUPON YOK","KUPON YOK","KUPON YOK"])
        except:
            try:
                element = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div/div/div[2]/div/div[1]/div[1]/span')))
            except:
                self.kuponcrawler()
            for i in range(1,30):
                try:        
                    title="Kupon adı="+self.driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div/div['+str(i)+']/div[1]/span').get_attribute('textContent') # başlık
                    initaldate=self.driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div/div['+str(i)+']/div[2]/div[1]/div[1]/div[1]').text #başlangıç tarihi
                    enddate=self.driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div/div['+str(i)+']/div[2]/div[1]/div[1]/div[2]').text #bitiş tarihi
                    try:
                        minshopprice=self.driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div/div['+str(i)+']/div[2]/div[1]/div[1]/div[3]').text #min alışveriş tutarı
                    except:
                        minshopprice="NO INFO"
                    givengift="KuponTutarı="+self.driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div/div['+str(i)+']/div[2]/div[1]/div[2]/div').text #kupon parası
                    try:
                        description="Açıklama="+self.driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div/div['+str(i)+']/div[2]/div[3]/div[2]/div').text
                    except Exception as E:
                        description=str(E)
                    self.verilist.append([title,initaldate,enddate,minshopprice,givengift,description])

                    
                except:
                    break
        

    
        self.excelwriter()

    def excelwriter(self):
        templist=[]
        for i in range(0,len(self.verilist)):
            templist.append(self.verilist[i][0]+","+self.verilist[i][1]+","+self.verilist[i][2]+","+self.verilist[i][3]+","+self.verilist[i][4]+","+self.verilist[i][5])
        templist.insert(0,self.username)
        templist.insert(1,self.passw)
        wb = load_workbook(str(self.lineEdit.text())+'.xlsx')
        ws = wb.worksheets[0]

        ws.append(templist)

        wb.save(str(self.lineEdit.text())+'.xlsx')

        self.bigstep+=len(self.verilist)
        
        self.logout()
    def logout(self):
        self.driver.get(str("https://m.trendyol.com/hesabim"))
        try:
            button = WebDriverWait(self.driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div[2]/div[2]/div/p')))
            button.click()
        except:
            return
    def couponcrawler(self,userlist,passlist):
        self.userlist=userlist
        self.passlist=passlist
        self.workbook = xlsxwriter.Workbook(str(self.lineEdit.text())+'.xlsx') 
        self.worksheet = self.workbook.add_worksheet()
        self.worksheet.write('A1', 'KullanıcıAdı') 
        self.worksheet.write('B1', 'Şifre') 
        self.worksheet.write('C1', 'Kupon1') 
        self.worksheet.write('D1', 'Kupon2') 
        self.worksheet.write('E1', 'Kupon3') 
        self.worksheet.write('F1', 'Kupon4')
        self.worksheet.write('G1', '...')
        self.workbook.close()
        
        for i in range(0,len(self.userlist)):
            self.username=self.userlist[i]
            if self.teksifre.isChecked():
                self.passw=self.passlist[0]
            else:
                for j in self.passlist:
                    self.passw=j
                    self.cond=self.login()
                    if self.cond==True:
                        break
            self.cond=self.login()
            if self.cond is True:
                self.kuponcrawler()
            self.verilist=[]
            self.step+=1
            self.couponprogressBar.setValue(self.step)
        

    def crawlorders(self):
        self.driver.get(str("https://m.trendyol.com/hesabim/siparislerim"))
        try:
            noorder = WebDriverWait(self.driver, 3).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div/div/div[2]/div/p')))
            self.verilist.append(["SIPARIS YOK","SIPARIS YOK","SIPARIS YOK","SIPARIS YOK","SIPARIS YOK"])
        except:
            try:
                element = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div[2]/div/div[1]/div[2]/p')))
            except:
                self.crawlorders()
            for i in range(1,30):
                try:                                    
                    date=self.driver.find_element_by_xpath('/html/body/main/div[2]/div/div[2]/div/div['+str(i)+']/div[1]/div/p').get_attribute('textContent') # başlık
                    ordercost=self.driver.find_element_by_xpath('/html/body/main/div[2]/div/div[2]/div/div['+str(i)+']/div[1]/div/div/p[2]').text #başlangıç tarihi
                    condition=self.driver.find_element_by_xpath('/html/body/main/div[2]/div/div[2]/div/div['+str(i)+']/div[2]/div[1]/p').text #bitiş tarihi
                    try:
                        explain=self.driver.find_element_by_xpath('/html/body/main/div[2]/div/div[2]/div/div['+str(i)+']/div[2]/div[3]/p').text #min alışveriş tutarı
                    except:
                        explain="NO INFO"
                    self.siparis_verilist.append([date,ordercost,condition,explain])
                except:
                    break
        self.orderexcellwriter()
    def orderexcellwriter(self):
        templist=[]
        for i in range(0,len(self.siparis_verilist)):
            templist.append(self.siparis_verilist[i][0]+","+self.siparis_verilist[i][1]+","+self.siparis_verilist[i][2]+","+self.siparis_verilist[i][3])
        templist.insert(0,self.username)
        templist.insert(1,self.passw)
        wb = load_workbook(str(self.lineEdit_siparis.text())+'.xlsx')
        ws = wb.worksheets[0]

        ws.append(templist)

        wb.save(str(self.lineEdit_siparis.text())+'.xlsx')

        self.bigstep_siparis+=len(self.siparis_verilist)
        
        self.logout()
    def getallorders(self,userlist,passlist):
        self.userlist=userlist
        self.passlist=passlist
        self.workbook = xlsxwriter.Workbook(str(self.lineEdit_siparis.text())+'.xlsx') 
        self.worksheet = self.workbook.add_worksheet()
        self.worksheet.write('A1', 'KullanıcıAdı') 
        self.worksheet.write('B1', 'Şifre') 
        self.worksheet.write('C1', 'Sipariş1') 
        self.worksheet.write('D1', 'Sipariş2') 
        self.worksheet.write('E1', 'Sipariş3') 
        self.worksheet.write('F1', '....')
        self.workbook.close()
        for i in range(0,len(self.userlist)):
            self.username=self.userlist[i]
            if self.teksifre_siparis.isChecked():
                self.passw=self.passlist[0]
            else:
                for j in self.passlist:
                    self.passw=j
                    self.cond=self.login()
                    if self.cond==True:
                        break
            self.cond=self.login()
            if self.cond is True:
                self.crawlorders()
            self.siparis_verilist=[]
            self.siparis_step+=1
            self.couponprogressBar_siparis.setValue(self.siparis_step)
        
    def pyqtchangepass(self):
        usertext = self.usernametextedit_chngpss.toPlainText()
        oldpasstext = self.passtextedit_chngoldpss.toPlainText()
        newpasstext = self.passtextedit_chngnewpss.toPlainText()
        self.userlist=usertext.split('\n')
        self.oldpasslist=oldpasstext.split('\n')
        self.newpasslist=newpasstext.split('\n')
        self.managingpassporcess()
        self.couponprogressBar_siparis.setMaximum(len(self.userlist))
    def changepassprocess(self):
        #time.sleep(500)
        try:
            control = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/header/div/div/div/a/img')))
        except:
            self.driver.close()
            self.driver = webdriver.Chrome(self.path, chrome_options=self.chrome_options)
            cond=self.login()
            if cond==True:
                self.changepassprocess()
            else:
                return
       
        self.driver.get(str("https://m.trendyol.com/SifreGuncelleme"))
        
        oldpasselement = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div[2]/form/div/div[1]/div/div/input')))

        oldpasselement.send_keys(self.passw)
        newpasselement=self.driver.find_element_by_xpath('/html/body/main/div/div[2]/form/div/div[2]/div/div/input') #newpass
        newpasselement.send_keys(self.newpassw)
        newpasselement2=self.driver.find_element_by_xpath('/html/body/main/div/div[2]/form/div/div[3]/div/div/input') #newpass
        newpasselement2.send_keys(self.newpassw)
        button= WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div[2]/form/button')))#changebutton

        button.click()
        button= WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div[2]/div/button')))
        button.click()

        
    def managingpassporcess(self):
        for i in range(0,len(self.userlist)):
            self.username=self.userlist[i]
            if self.teksifre_chngoldpss.isChecked():
                self.passw=self.oldpasslist[0]
            else:
                self.passw=self.oldpasslist[i]
            if self.teksifre_chngnewpss.isChecked():
                self.newpassw=self.newpasslist[0]
            else:
                self.newpassw=self.newpasslist[i]
            self.driver.close()
            self.driver = webdriver.Chrome(self.path, chrome_options=self.chrome_options)
            self.cond=self.login()
            if self.cond is True:
                self.changepassprocess()
            self.chngpss_step+=1
            self.couponprogressBar_chngpss.setValue(self.chngpss_step)






if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
    
     
