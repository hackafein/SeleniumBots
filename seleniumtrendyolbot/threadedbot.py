import os, time, random
from selenium import webdriver
import xlsxwriter
from selenium.webdriver.chrome.options import Options
import timeit
import re
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from PyQt5 import QtCore, QtGui, QtWidgets

from openpyxl import Workbook
from openpyxl import load_workbook
import threading

class Ui_MainWindow(QtWidgets.QWidget):


    def setupUi(self, MainWindow):

        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 600)
        self.kuponprogresyuzde=0
        self.siparisprogresyuzde=0
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setGeometry(QtCore.QRect(0, 0, 791, 551))
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.pushButton = QtWidgets.QPushButton(self.tab)
        self.pushButton.setGeometry(QtCore.QRect(600, 450, 113, 32))
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.couponcrawl)
        self.couponprogressBar = QtWidgets.QProgressBar(self.tab)
        self.couponprogressBar.setGeometry(QtCore.QRect(200, 450, 391, 21))
        self.couponprogressBar.setProperty("value", 0)
        self.couponprogressBar.setObjectName("couponprogressBar")
        self.usernametextedit = QtWidgets.QPlainTextEdit(self.tab)
        self.usernametextedit.setGeometry(QtCore.QRect(50, 70, 311, 261))
        self.usernametextedit.setObjectName("usernametextedit")
        self.passtextedit = QtWidgets.QPlainTextEdit(self.tab)
        self.passtextedit.setGeometry(QtCore.QRect(440, 70, 321, 261))
        self.passtextedit.setObjectName("passtextedit")
        self.label = QtWidgets.QLabel(self.tab)
        self.label.setGeometry(QtCore.QRect(150, 40, 101, 21))
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.tab)
        self.label_2.setGeometry(QtCore.QRect(580, 40, 101, 21))
        self.label_2.setObjectName("label_2")
        self.lineEdit = QtWidgets.QLineEdit(self.tab)
        self.lineEdit.setGeometry(QtCore.QRect(340, 380, 113, 21))
        self.lineEdit.setObjectName("lineEdit")
        self.label_3 = QtWidgets.QLabel(self.tab)
        self.label_3.setGeometry(QtCore.QRect(230, 380, 111, 20))
        self.label_3.setObjectName("label_3")
        self.teksifre = QtWidgets.QCheckBox(self.tab)
        self.teksifre.setGeometry(QtCore.QRect(600, 340, 161, 21))
        self.teksifre.setObjectName("teksifre")
        self.browserspinBox1 = QtWidgets.QSpinBox(self.tab)
        self.browserspinBox1.setGeometry(QtCore.QRect(140, 380, 81, 24))
        self.browserspinBox1.setMinimum(1)
        self.browserspinBox1.setObjectName("browserspinBox1")
        self.browserlabel1 = QtWidgets.QLabel(self.tab)
        self.browserlabel1.setGeometry(QtCore.QRect(40, 380, 111, 21))
        self.browserlabel1.setObjectName("browserlabel1")
        self.notdefterikullanicikupon = QtWidgets.QPushButton(self.tab)
        self.notdefterikullanicikupon.setGeometry(QtCore.QRect(50, 330, 311, 32))
        self.notdefterikullanicikupon.setObjectName("notdefterikullanicikupon")
        self.notdefterikullanicikupon.clicked.connect(self.filediyalogkuponusername)
        self.notdefterisifrekupon = QtWidgets.QPushButton(self.tab)
        self.notdefterisifrekupon.setGeometry(QtCore.QRect(440, 330, 161, 32))
        self.notdefterisifrekupon.clicked.connect(self.filediyalogkuponpassword)
        self.notdefterisifrekupon.setObjectName("notdefterisifrekupon")
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tabWidget.addTab(self.tab, "")
        self.siparislericek = QtWidgets.QWidget()
        self.siparislericek.setObjectName("siparislericek")
        self.lineEdit_siparis = QtWidgets.QLineEdit(self.siparislericek)
        self.lineEdit_siparis.setGeometry(QtCore.QRect(320, 370, 113, 21))
        self.lineEdit_siparis.setObjectName("lineEdit_siparis")
        self.exellabelsiparis = QtWidgets.QLabel(self.siparislericek)
        self.exellabelsiparis.setGeometry(QtCore.QRect(210, 370, 111, 20))
        self.exellabelsiparis.setObjectName("exellabelsiparis")
        self.sipariskullaniciadilabel = QtWidgets.QLabel(self.siparislericek)
        self.sipariskullaniciadilabel.setGeometry(QtCore.QRect(130, 30, 101, 21))
        self.sipariskullaniciadilabel.setObjectName("sipariskullaniciadilabel")
        self.siparissifrelabel = QtWidgets.QLabel(self.siparislericek)
        self.siparissifrelabel.setGeometry(QtCore.QRect(560, 30, 101, 21))
        self.siparissifrelabel.setObjectName("siparissifrelabel")
        self.pushButton_siparis = QtWidgets.QPushButton(self.siparislericek)
        self.pushButton_siparis.setGeometry(QtCore.QRect(580, 440, 113, 32))
        self.pushButton_siparis.setObjectName("pushButton_siparis")
        self.pushButton_siparis.clicked.connect(self.getorderspyqt)
        self.passtextedit_siparis = QtWidgets.QPlainTextEdit(self.siparislericek)
        self.passtextedit_siparis.setGeometry(QtCore.QRect(420, 60, 321, 261))
        self.passtextedit_siparis.setObjectName("passtextedit_siparis")
        self.couponprogressBar_siparis = QtWidgets.QProgressBar(self.siparislericek)
        self.couponprogressBar_siparis.setGeometry(QtCore.QRect(180, 440, 391, 21))
        self.couponprogressBar_siparis.setProperty("value", 0)
        self.couponprogressBar_siparis.setObjectName("couponprogressBar_siparis")
        self.usernametextedit_siparis = QtWidgets.QPlainTextEdit(self.siparislericek)
        self.usernametextedit_siparis.setGeometry(QtCore.QRect(30, 60, 311, 261))
        self.usernametextedit_siparis.setObjectName("usernametextedit_siparis")
        self.teksifre_siparis = QtWidgets.QCheckBox(self.siparislericek)
        self.teksifre_siparis.setGeometry(QtCore.QRect(580, 330, 161, 21))
        self.teksifre_siparis.setObjectName("teksifre_siparis")
        self.notdefterikullanicisiparis = QtWidgets.QPushButton(self.siparislericek)
        self.notdefterikullanicisiparis.setGeometry(QtCore.QRect(20, 320, 311, 32))
        self.notdefterikullanicisiparis.setObjectName("notdefterikullanicisiparis")
        self.notdefterikullanicisiparis.clicked.connect(self.filediyalogsiparisusername)
        self.browserspinBox2 = QtWidgets.QSpinBox(self.siparislericek)
        self.browserspinBox2.setGeometry(QtCore.QRect(120, 370, 81, 24))
        self.browserspinBox2.setMinimum(1)
        self.browserspinBox2.setObjectName("browserspinBox2")
        self.browserlabel2 = QtWidgets.QLabel(self.siparislericek)
        self.browserlabel2.setGeometry(QtCore.QRect(20, 370, 111, 21))
        self.browserlabel2.setObjectName("browserlabel2")
        self.notdefterisifresiparis = QtWidgets.QPushButton(self.siparislericek)
        self.notdefterisifresiparis.setGeometry(QtCore.QRect(410, 320, 161, 32))
        self.notdefterisifresiparis.setObjectName("notdefterisifresiparis")
        self.notdefterisifresiparis.clicked.connect(self.filediyalogsiparispassword)
        self.tabWidget.addTab(self.siparislericek, "")

        self.tab_3 = QtWidgets.QWidget()
        self.tab_3.setObjectName("tab_3")
        self.label_4 = QtWidgets.QLabel(self.tab_3)
        self.label_4.setGeometry(QtCore.QRect(100, 20, 101, 21))
        self.label_4.setObjectName("label_4")
        self.usernametextedit_chngpss = QtWidgets.QPlainTextEdit(self.tab_3)
        self.usernametextedit_chngpss.setGeometry(QtCore.QRect(40, 50, 251, 261))
        self.usernametextedit_chngpss.setObjectName("usernametextedit_chngpss")
        self.label_5 = QtWidgets.QLabel(self.tab_3)
        self.label_5.setGeometry(QtCore.QRect(360, 20, 101, 21))
        self.label_5.setObjectName("label_5")
        self.passtextedit_chngoldpss = QtWidgets.QPlainTextEdit(self.tab_3)
        self.passtextedit_chngoldpss.setGeometry(QtCore.QRect(310, 50, 211, 261))
        self.passtextedit_chngoldpss.setObjectName("passtextedit_chngoldpss")
        self.teksifre_chngoldpss = QtWidgets.QCheckBox(self.tab_3)
        self.teksifre_chngoldpss.setGeometry(QtCore.QRect(340, 320, 161, 21))
        self.teksifre_chngoldpss.setObjectName("teksifre_chngoldpss")
        self.pushButton_chngpss = QtWidgets.QPushButton(self.tab_3)
        self.pushButton_chngpss.setGeometry(QtCore.QRect(550, 450, 141, 41))
        self.pushButton_chngpss.setObjectName("pushButton_chngpss")
        self.pushButton_chngpss.clicked.connect(self.pyqtchangepass)
        self.couponprogressBar_chngpss = QtWidgets.QProgressBar(self.tab_3)
        self.couponprogressBar_chngpss.setGeometry(QtCore.QRect(140, 460, 391, 21))
        self.couponprogressBar_chngpss.setProperty("value", 0)
        self.couponprogressBar_chngpss.setObjectName("couponprogressBar_chngpss")
        self.teksifre_chngnewpss = QtWidgets.QCheckBox(self.tab_3)
        self.teksifre_chngnewpss.setGeometry(QtCore.QRect(580, 320, 161, 21))
        self.teksifre_chngnewpss.setObjectName("teksifre_chngnewpss")
        self.passtextedit_chngnewpss = QtWidgets.QPlainTextEdit(self.tab_3)
        self.passtextedit_chngnewpss.setGeometry(QtCore.QRect(560, 50, 211, 261))
        self.passtextedit_chngnewpss.setObjectName("passtextedit_chngnewpss")
        self.label_6 = QtWidgets.QLabel(self.tab_3)
        self.label_6.setGeometry(QtCore.QRect(610, 20, 101, 21))
        self.label_6.setObjectName("label_6")
        self.tabWidget.addTab(self.tab_3, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.label_7 = QtWidgets.QLabel(self.tab_2)
        self.label_7.setGeometry(QtCore.QRect(520, 10, 101, 21))
        self.label_7.setObjectName("label_7")
 
        self.label_9 = QtWidgets.QLabel(self.tab_2)
        self.label_9.setGeometry(QtCore.QRect(40, 340, 91, 16))
        self.label_9.setObjectName("label_9")
        self.label_10 = QtWidgets.QLabel(self.tab_2)
        self.label_10.setGeometry(QtCore.QRect(40, 370, 91, 16))
        self.label_10.setObjectName("label_10")
        self.label_11 = QtWidgets.QLabel(self.tab_2)
        self.label_11.setGeometry(QtCore.QRect(220, 370, 61, 16))
        self.label_11.setObjectName("label_11")
        self.label_12 = QtWidgets.QLabel(self.tab_2)
        self.label_12.setGeometry(QtCore.QRect(40, 410, 101, 16))
        self.label_12.setObjectName("label_12")
        self.label_16 = QtWidgets.QLabel(self.tab_2)
        self.label_16.setGeometry(QtCore.QRect(30, 490, 101, 16))
        self.label_16.setObjectName("label_16")
        self.tabWidget.addTab(self.tab_2, "")

        self.Adres_Kullaniciadiyukle = QtWidgets.QPushButton(self.tab_2)
        self.Adres_Kullaniciadiyukle.setGeometry(QtCore.QRect(30, 50, 261, 32))
        self.Adres_Kullaniciadiyukle.setObjectName("Adres_Kullaniciadiyukle")
        self.Adres_SifreYukle = QtWidgets.QPushButton(self.tab_2)
        self.Adres_SifreYukle.setGeometry(QtCore.QRect(30, 100, 261, 32))
        self.Adres_SifreYukle.setObjectName("Adres_SifreYukle")
        self.Adres_Adyukle = QtWidgets.QPushButton(self.tab_2)
        self.Adres_Adyukle.setGeometry(QtCore.QRect(30, 150, 261, 32))
        self.Adres_Adyukle.setObjectName("Adres_Adyukle")
        self.Adres_SoyadYukle = QtWidgets.QPushButton(self.tab_2)
        self.Adres_SoyadYukle.setGeometry(QtCore.QRect(30, 200, 261, 32))
        self.Adres_SoyadYukle.setObjectName("Adres_SoyadYukle")
        self.Adres_CeptelYukle = QtWidgets.QPushButton(self.tab_2)
        self.Adres_CeptelYukle.setGeometry(QtCore.QRect(30, 250, 261, 32))
        self.Adres_CeptelYukle.setObjectName("Adres_CeptelYukle")
        self.Adres_IlYukle = QtWidgets.QPushButton(self.tab_2)
        self.Adres_IlYukle.setGeometry(QtCore.QRect(30, 300, 261, 32))
        self.Adres_IlYukle.setObjectName("Adres_IlYukle")
        self.Adres_IlceYukle = QtWidgets.QPushButton(self.tab_2)
        self.Adres_IlceYukle.setGeometry(QtCore.QRect(30, 350, 261, 32))
        self.Adres_IlceYukle.setObjectName("Adres_IlceYukle")
        self.Adres_MahalleYukle = QtWidgets.QPushButton(self.tab_2)
        self.Adres_MahalleYukle.setGeometry(QtCore.QRect(30, 400, 261, 32))
        self.Adres_MahalleYukle.setObjectName("Adres_MahalleYukle")
        self.Adres_AdresYukle = QtWidgets.QPushButton(self.tab_2)
        self.Adres_AdresYukle.setGeometry(QtCore.QRect(30, 450, 261, 32))
        self.Adres_AdresYukle.setObjectName("Adres_AdresYukle")
        self.Adres_KullaniciAdiLabel = QtWidgets.QLabel(self.tab_2)
        self.Adres_KullaniciAdiLabel.setGeometry(QtCore.QRect(330, 60, 101, 16))
        self.Adres_KullaniciAdiLabel.setObjectName("Adres_KullaniciAdiLabel")
        self.Adres_SifreLabel = QtWidgets.QLabel(self.tab_2)
        self.Adres_SifreLabel.setGeometry(QtCore.QRect(330, 110, 101, 16))
        self.Adres_SifreLabel.setObjectName("Adres_SifreLabel")
        self.Adres_AdLabel = QtWidgets.QLabel(self.tab_2)
        self.Adres_AdLabel.setGeometry(QtCore.QRect(330, 160, 101, 16))
        self.Adres_AdLabel.setObjectName("Adres_AdLabel")
        self.Adres_SoyadLabel = QtWidgets.QLabel(self.tab_2)
        self.Adres_SoyadLabel.setGeometry(QtCore.QRect(330, 210, 101, 16))
        self.Adres_SoyadLabel.setObjectName("Adres_SoyadLabel")
        self.Adres_CeptelLabel = QtWidgets.QLabel(self.tab_2)
        self.Adres_CeptelLabel.setGeometry(QtCore.QRect(330, 260, 101, 16))
        self.Adres_CeptelLabel.setObjectName("Adres_CeptelLabel")
        self.Adres_IlLabel = QtWidgets.QLabel(self.tab_2)
        self.Adres_IlLabel.setGeometry(QtCore.QRect(330, 310, 101, 16))
        self.Adres_IlLabel.setObjectName("Adres_IlLabel")
        self.Adres_IlceLabel = QtWidgets.QLabel(self.tab_2)
        self.Adres_IlceLabel.setGeometry(QtCore.QRect(330, 360, 101, 16))
        self.Adres_IlceLabel.setObjectName("Adres_IlceLabel")
        self.Adres_MahalleLabel = QtWidgets.QLabel(self.tab_2)
        self.Adres_MahalleLabel.setGeometry(QtCore.QRect(330, 410, 101, 16))
        self.Adres_MahalleLabel.setObjectName("Adres_MahalleLabel")
        self.Adres_AdresLabel = QtWidgets.QLabel(self.tab_2)
        self.Adres_AdresLabel.setGeometry(QtCore.QRect(330, 460, 101, 16))
        self.Adres_AdresLabel.setObjectName("Adres_AdresLabel")
        self.line = QtWidgets.QFrame(self.tab_2)
        self.line.setGeometry(QtCore.QRect(0, 80, 781, 16))
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.line_2 = QtWidgets.QFrame(self.tab_2)
        self.line_2.setGeometry(QtCore.QRect(0, 130, 781, 16))
        self.line_2.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")
        self.line_3 = QtWidgets.QFrame(self.tab_2)
        self.line_3.setGeometry(QtCore.QRect(0, 180, 781, 16))
        self.line_3.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_3.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_3.setObjectName("line_3")
        self.line_4 = QtWidgets.QFrame(self.tab_2)
        self.line_4.setGeometry(QtCore.QRect(10, 230, 781, 16))
        self.line_4.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_4.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_4.setObjectName("line_4")
        self.line_5 = QtWidgets.QFrame(self.tab_2)
        self.line_5.setGeometry(QtCore.QRect(0, 280, 781, 16))
        self.line_5.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_5.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_5.setObjectName("line_5")
        self.line_6 = QtWidgets.QFrame(self.tab_2)
        self.line_6.setGeometry(QtCore.QRect(0, 330, 781, 16))
        self.line_6.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_6.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_6.setObjectName("line_6")
        self.line_7 = QtWidgets.QFrame(self.tab_2)
        self.line_7.setGeometry(QtCore.QRect(0, 380, 781, 16))
        self.line_7.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_7.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_7.setObjectName("line_7")
        self.line_8 = QtWidgets.QFrame(self.tab_2)
        self.line_8.setGeometry(QtCore.QRect(-10, 430, 801, 20))
        self.line_8.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_8.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_8.setObjectName("line_8")
        self.line_9 = QtWidgets.QFrame(self.tab_2)
        self.line_9.setGeometry(QtCore.QRect(0, 30, 781, 16))
        self.line_9.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_9.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_9.setObjectName("line_9")
        self.Adres_Adresleriekle = QtWidgets.QPushButton(self.tab_2)
        self.Adres_Adresleriekle.setGeometry(QtCore.QRect(590, 480, 191, 41))
        self.Adres_Adresleriekle.setObjectName("Adres_Adresleriekle")
        self.Adres_Progressbar = QtWidgets.QProgressBar(self.tab_2)
        self.Adres_Progressbar.setGeometry(QtCore.QRect(40, 490, 551, 21))
        self.Adres_Progressbar.setProperty("value", 0)
        self.Adres_Progressbar.setObjectName("Adres_Progressbar")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 24))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        
        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Trendyol Crawler"))
        self.pushButton.setText(_translate("MainWindow", "Kuponları Çek"))
        self.label.setText(_translate("MainWindow", "Kullanıcı adları"))
        self.label_2.setText(_translate("MainWindow", "Şifreler"))
        self.label_3.setText(_translate("MainWindow", "Excel dosya adı ="))
        self.teksifre.setText(_translate("MainWindow", "Hepsi için tek şifre"))
        self.browserlabel1.setText(_translate("MainWindow", "Browser sayısı ="))
        self.notdefterikullanicikupon.setText(_translate("MainWindow", "Not defterinden Yükle"))
        self.notdefterisifrekupon.setText(_translate("MainWindow", "Not defterinden Yükle"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("MainWindow", "Kuponları Çek"))
        self.exellabelsiparis.setText(_translate("MainWindow", "Excel dosya adı ="))
        self.sipariskullaniciadilabel.setText(_translate("MainWindow", "Kullanıcı adları"))
        self.siparissifrelabel.setText(_translate("MainWindow", "Şifreler"))
        self.pushButton_siparis.setText(_translate("MainWindow", "Siparişleri Çek"))
        self.teksifre_siparis.setText(_translate("MainWindow", "Hepsi için tek şifre"))
        self.notdefterikullanicisiparis.setText(_translate("MainWindow", "Not defterinden Yükle"))
        self.browserlabel2.setText(_translate("MainWindow", "Browser sayısı ="))
        self.notdefterisifresiparis.setText(_translate("MainWindow", "Not defterinden Yükle"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.siparislericek), _translate("MainWindow", "Siparişleri Çek"))
        self.label_4.setText(_translate("MainWindow", "Kullanıcı adları"))
        self.label_5.setText(_translate("MainWindow", "Eski Şifreler"))
        self.teksifre_chngoldpss.setText(_translate("MainWindow", "Hepsi için tek şifre"))
        self.pushButton_chngpss.setText(_translate("MainWindow", "Şifreleri Değiştir"))
        self.teksifre_chngnewpss.setText(_translate("MainWindow", "Hepsi için tek şifre koy"))
        self.label_6.setText(_translate("MainWindow", "Yeni Şifreler"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_3), _translate("MainWindow", "Şifre Değiştir"))

        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("MainWindow", "Adres Ekle"))
        self.Adres_Kullaniciadiyukle.setText(_translate("MainWindow", "Kullanıcı Adı Yükle"))
        self.Adres_SifreYukle.setText(_translate("MainWindow", "Şifre Yükle"))
        self.Adres_Adyukle.setText(_translate("MainWindow", "Ad Yükle"))
        self.Adres_SoyadYukle.setText(_translate("MainWindow", "Soyad Yükle"))
        self.Adres_CeptelYukle.setText(_translate("MainWindow", "Ceptel Yükle"))
        self.Adres_IlYukle.setText(_translate("MainWindow", "İl Yükle"))
        self.Adres_IlceYukle.setText(_translate("MainWindow", "İlçe Yükle"))
        self.Adres_MahalleYukle.setText(_translate("MainWindow", "Mahalle Yükle"))
        self.Adres_AdresYukle.setText(_translate("MainWindow", "Adres Yükle"))
        self.Adres_KullaniciAdiLabel.setText(_translate("MainWindow", "Yüklenmedi"))
        self.Adres_SifreLabel.setText(_translate("MainWindow", "Yüklenmedi"))
        self.Adres_AdLabel.setText(_translate("MainWindow", "Yüklenmedi"))
        self.Adres_SoyadLabel.setText(_translate("MainWindow", "Yüklenmedi"))
        self.Adres_CeptelLabel.setText(_translate("MainWindow", "Yüklenmedi"))
        self.Adres_IlLabel.setText(_translate("MainWindow", "Yüklenmedi"))
        self.Adres_IlceLabel.setText(_translate("MainWindow", "Yüklenmedi"))
        self.Adres_MahalleLabel.setText(_translate("MainWindow", "Yüklenmedi"))
        self.Adres_AdresLabel.setText(_translate("MainWindow", "Yüklenmedi"))
        self.Adres_Adresleriekle.setText(_translate("MainWindow", "Adresleri Ekle"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("MainWindow", "Adres Ekle"))
        self.couponprogressBar.setValue(self.kuponprogresyuzde)
        self.couponprogressBar_siparis.setValue(self.siparisprogresyuzde)
    def filediyalogkuponusername(self):
        data_path=QtWidgets.QFileDialog.getOpenFileName(self,'Txt yükle',"",'*.txt',)
        f = open(data_path[0], "r")
        for x in f:
            self.usernametextedit.appendPlainText(x.replace("\n",""))
    def filediyalogkuponpassword(self):
        data_path=QtWidgets.QFileDialog.getOpenFileName(self,'Txt yükle',"",'*.txt',)
        f = open(data_path[0], "r")
        for x in f:
            self.passtextedit.appendPlainText(x.replace("\n",""))
    def filediyalogsiparisusername(self):
        data_path=QtWidgets.QFileDialog.getOpenFileName(self,'Txt yükle',"",'*.txt',)
        f = open(data_path[0], "r")
        for x in f:
            self.usernametextedit_siparis.appendPlainText(x.replace("\n",""))
    def filediyalogsiparispassword(self):
        data_path=QtWidgets.QFileDialog.getOpenFileName(self,'Txt yükle',"",'*.txt',)
        f = open(data_path[0], "r")
        for x in f:
            self.passtextedit_siparis.appendPlainText(x.replace("\n",""))
    def chunkIt(self,seq, num):
        avg = len(seq) / float(num)
        out = []
        last = 0.0

        while last < len(seq):
            out.append(seq[int(last):int(last + avg)])
            last += avg

        return out
    def couponcrawl(self):
        usertext = self.usernametextedit.toPlainText()
        passtext = self.passtextedit.toPlainText()
        self.userlist=usertext.split('\n')
        self.passlist=passtext.split('\n')

        threadcount=int(self.browserspinBox1.value())

        self.userlistnew=self.chunkIt(self.userlist,threadcount)

        self.passlistnew=self.passlist

        worker=[]
        for i in range(0,threadcount):
            worker.append(threading.Thread(target=TrendyolSelenium, args = (i+1,"KUPON",self.userlistnew[i],self.passlistnew,self, )))
            
        for i in worker:
            i.start()

        #self.couponcrawler(self.userlist1,self.passlist1)
        self.couponprogressBar.setMaximum(len(self.userlist))


    def getorderspyqt(self):
        usertext = self.usernametextedit_siparis.toPlainText()
        passtext = self.passtextedit_siparis.toPlainText()
        self.userlist=usertext.split('\n')
        self.passlist=passtext.split('\n')
        #self.getallorders(self.userlist,self.passlist)
        
        threadcount=int(self.browserspinBox2.value())

        self.userlistnew=self.chunkIt(self.userlist,threadcount)

        self.passlistnew=self.passlist

        worker=[]
        for i in range(0,threadcount):
            worker.append(threading.Thread(target=TrendyolSelenium, args = (i+1,"ORDER",self.userlistnew[i],self.passlistnew,self, )))
            
        for i in worker:
            i.start()

        #self.couponcrawler(self.userlist1,self.passlist1)
        self.couponprogressBar_siparis.setMaximum(len(self.userlist))




    def pyqtchangepass(self):
        usertext = self.usernametextedit_chngpss.toPlainText()
        oldpasstext = self.passtextedit_chngoldpss.toPlainText()
        newpasstext = self.passtextedit_chngnewpss.toPlainText()
        self.userlist=usertext.split('\n')
        self.oldpasslist=oldpasstext.split('\n')
        self.newpasslist=newpasstext.split('\n')
        #self.couponprogressBar_siparis.setMaximum(len(self.userlist))
        threadcount=1
        self.userlistnew=self.userlist
        self.passlistnew=[self.newpasslist,self.oldpasslist]
        worker=[]
        for i in range(0,threadcount):
            worker.append(threading.Thread(target=TrendyolSelenium, args = (i+1,"PASSWORD",self.userlistnew,self.passlistnew,self, )))
        for i in worker:
            i.start()
        #self.couponcrawler(self.userlist1,self.passlist1)
        #self.couponprogressBar_siparis.setMaximum(len(self.userlist))

class TrendyolSelenium(threading.Thread):
    def __init__(self, threadID,action, username,password,pyqt):
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.pyqt=pyqt
        self.action=action
        self.path = f'{os.getcwd()}/chromedriver'
        self.chrome_options = Options()
        mobile_emulation = {
        "deviceMetrics": {"width": 360, "height": 640, "pixelRatio": 3.0},
        "userAgent": "Mozilla/5.0 (Linux; Android 4.2.1; en-us; Nexus 5 Build/JOP40D) AppleWebKit/535.19 (KHTML, like Gecko) Chrome/18.0.1025.166 Mobile Safari/535.19"}
        #self.chrome_options.add_argument("--headless")
        self.chrome_options.add_experimental_option("mobileEmulation", mobile_emulation)
        
        self.verilist=[]
        self.userlist=[]
        self.passlist=[]
        self.step=0
        self.siparis_step=0
        self.siparis_verilist=[]
        self.bigstep=0
        self.bigstep_siparis=0
        self.chngpss_step=0
        self.username=""
        self.passw=""
        self.cond=True
        if action=="KUPON":
            self.couponcrawler(username,password)
        elif action=="ORDER":
            
            self.getallorders(username,password)
        elif action=="PASSWORD":
            self.userlist=username
            newpass=password[0]
            oldpass=password[1]
            
            self.managingpassporcess(username,newpass,oldpass)


    def login(self):
        try:
            self.driver.get(str("https://m.trendyol.com/giris?cb=%2F"))

            userelement= WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div[2]/form/div[1]/input')))

            userelement.send_keys(self.username)
            
            passelement=self.driver.find_element_by_xpath('/html/body/main/div/div[2]/form/div[2]/div/div/input')
            passelement.send_keys(self.passw)
            button= WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div[2]/form/button')))
            button.click()
        
            time.sleep(4)
            try:
                self.driver.find_element_by_xpath("/html/body/main/div/div[2]/span[2]")
                self.verilist.append(["GIRIS BASARISIZ","GIRIS BASARISIZ","GIRIS BASARISIZ","GIRIS BASARISIZ","GIRIS BASARISIZ"])
                #self.kuponcrawler()
                return False
            except:
                return True
        except:
            self.login()
            
            
    def kuponcrawler(self):
        self.driver.get(str("https://m.trendyol.com/hesabim/indirim-kuponlari"))
        try:
            trial = WebDriverWait(self.driver, 3).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div/div/div[1]/div[2]/p')))
        except:
            self.driver.close()
            self.driver = webdriver.Chrome(self.path, chrome_options=self.chrome_options)
            cond=self.login()
            if cond==True:
                self.kuponcrawler()
        try:
            nocupon = WebDriverWait(self.driver, 3).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div/div/div[2]/div/p')))
            self.verilist.append(["KUPON YOK","KUPON YOK","KUPON YOK","KUPON YOK","KUPON YOK"])
        except:
            try:
                element = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div/div/div[2]/div/div[1]/div[1]/span')))
            except:
                self.kuponcrawler()
            for i in range(1,30):
                try:        
                    title="Kupon adı="+self.driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div/div['+str(i)+']/div[1]/span').get_attribute('textContent') # başlık
                    initaldate=self.driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div/div['+str(i)+']/div[2]/div[1]/div[1]/div[1]').text #başlangıç tarihi
                    enddate=self.driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div/div['+str(i)+']/div[2]/div[1]/div[1]/div[2]').text #bitiş tarihi
                    try:
                        minshopprice=self.driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div/div['+str(i)+']/div[2]/div[1]/div[1]/div[3]').text #min alışveriş tutarı
                    except:
                        minshopprice="NO INFO"
                    givengift="KuponTutarı="+self.driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div/div['+str(i)+']/div[2]/div[1]/div[2]/div').text #kupon parası
                    try:
                        description="Açıklama="+self.driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div/div['+str(i)+']/div[2]/div[3]/div[2]/div').text
                    except Exception as E:
                        description=str(E)
                    self.verilist.append([title,initaldate,enddate,minshopprice,givengift,description])

                    
                except:
                    break
        

    
        self.excelwriter()

    def excelwriter(self):
        templist=[]
        for i in range(0,len(self.verilist)):
            templist.append(self.verilist[i][0]+","+self.verilist[i][1]+","+self.verilist[i][2]+","+self.verilist[i][3]+","+self.verilist[i][4]+","+self.verilist[i][5])
        templist.insert(0,self.username)
        templist.insert(1,self.passw)
        wb = load_workbook(str(self.pyqt.lineEdit.text())+'.xlsx')
        ws = wb.worksheets[0]

        ws.append(templist)

        wb.save(str(self.pyqt.lineEdit.text())+'.xlsx')

        self.bigstep+=len(self.verilist)
        
        self.logout()
    def logout(self):
        self.driver.get(str("https://m.trendyol.com/hesabim"))
        try:
            button = WebDriverWait(self.driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div[2]/div[2]/div/p')))
            button.click()
        except:
            return
    def couponcrawler(self,userlist,passlist):
        self.driver = webdriver.Chrome(self.path, chrome_options=self.chrome_options)
        self.userlist=userlist
        self.passlist=passlist
        self.workbook = xlsxwriter.Workbook(str(self.pyqt.lineEdit.text())+'.xlsx') 
        self.worksheet = self.workbook.add_worksheet()
        self.worksheet.write('A1', 'KullanıcıAdı') 
        self.worksheet.write('B1', 'Şifre') 
        self.worksheet.write('C1', 'Kupon1') 
        self.worksheet.write('D1', 'Kupon2') 
        self.worksheet.write('E1', 'Kupon3') 
        self.worksheet.write('F1', 'Kupon4')
        self.worksheet.write('G1', '...')
        self.workbook.close()
        
        for i in range(0,len(self.userlist)):
            self.username=self.userlist[i]
            if self.pyqt.teksifre.isChecked():
                self.passw=self.passlist[0]
            else:
                for j in self.passlist:
                    self.passw=j
                    self.cond=self.login()
                    if self.cond==True:
                        break
            self.cond=self.login()
            if self.cond is True:
                self.kuponcrawler()
            self.verilist=[]
            self.step+=1
            self.pyqt.kuponprogresyuzde+=1
            if self.step>=len(self.userlist):
                break
            
        self.pyqt.couponprogressBar.setValue(self.pyqt.kuponprogresyuzde)
        self.driver.close()
        sys.exit()
    def crawlorders(self):
        self.driver.get(str("https://m.trendyol.com/hesabim/siparislerim"))        
        try:
            trial = WebDriverWait(self.driver, 3).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div[2]/div/div[1]/div[2]/p')))
        except:
            self.driver.close()
            self.driver = webdriver.Chrome(self.path, chrome_options=self.chrome_options)
            cond=self.login()
            if cond==True:
                self.crawlorders()
        try:
            noorder = WebDriverWait(self.driver, 3).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div/div/div[2]/div/p')))
            self.verilist.append(["SIPARIS YOK","SIPARIS YOK","SIPARIS YOK","SIPARIS YOK","SIPARIS YOK"])
        except:
            try:
                element = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div[2]/div/div[1]/div[2]/p')))
            except:
                self.crawlorders()
            for i in range(1,30):
                try:                                    
                    date=self.driver.find_element_by_xpath('/html/body/main/div[2]/div/div[2]/div/div['+str(i)+']/div[1]/div/p').get_attribute('textContent') # başlık
                    ordercost=self.driver.find_element_by_xpath('/html/body/main/div[2]/div/div[2]/div/div['+str(i)+']/div[1]/div/div/p[2]').text #başlangıç tarihi
                    condition=self.driver.find_element_by_xpath('/html/body/main/div[2]/div/div[2]/div/div['+str(i)+']/div[2]/div[1]/p').text #bitiş tarihi
                    try:
                        explain=self.driver.find_element_by_xpath('/html/body/main/div[2]/div/div[2]/div/div['+str(i)+']/div[2]/div[3]/p').text #min alışveriş tutarı
                    except:
                        explain="NO INFO"
                    self.siparis_verilist.append([date,ordercost,condition,explain])
                except:
                    break
        self.orderexcellwriter()
    def orderexcellwriter(self):
        templist=[]
        for i in range(0,len(self.siparis_verilist)):
            templist.append(self.siparis_verilist[i][0]+","+self.siparis_verilist[i][1]+","+self.siparis_verilist[i][2]+","+self.siparis_verilist[i][3])
        templist.insert(0,self.username)
        templist.insert(1,self.passw)
        wb = load_workbook(str(self.pyqt.lineEdit_siparis.text())+'.xlsx')
        ws = wb.worksheets[0]

        ws.append(templist)

        wb.save(str(self.pyqt.lineEdit_siparis.text())+'.xlsx')

        self.bigstep_siparis+=len(self.siparis_verilist)
        
        self.logout()
    def getallorders(self,userlist,passlist):
        self.driver = webdriver.Chrome(self.path, chrome_options=self.chrome_options)
        self.userlist=userlist
        self.passlist=passlist
        self.workbook = xlsxwriter.Workbook(str(self.pyqt.lineEdit_siparis.text())+'.xlsx') 
        self.worksheet = self.workbook.add_worksheet()
        self.worksheet.write('A1', 'KullanıcıAdı') 
        self.worksheet.write('B1', 'Şifre') 
        self.worksheet.write('C1', 'Sipariş1') 
        self.worksheet.write('D1', 'Sipariş2') 
        self.worksheet.write('E1', 'Sipariş3') 
        self.worksheet.write('F1', '....')
        self.workbook.close()
        for i in range(0,len(self.userlist)):
            self.username=self.userlist[i]
            if self.pyqt.teksifre_siparis.isChecked():
                self.passw=self.passlist[0]
            else:
                for j in self.passlist:
                    self.passw=j
                    self.cond=self.login()
                    if self.cond==True:
                        break
            self.cond=self.login()
            if self.cond is True:
                self.crawlorders()
            self.siparis_verilist=[]
            self.siparis_step+=1
            self.pyqt.siparisprogresyuzde+=1
            if self.siparis_step>=len(self.userlist):
                break
            
        self.pyqt.couponprogressBar_siparis.setValue(self.pyqt.siparisprogresyuzde)
        self.driver.close()
        sys.exit()

    def changepassprocess(self):
        #time.sleep(500)
        try:
            control = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/header/div/div/div/a/img')))
        except:
            self.driver.close()
            self.driver = webdriver.Chrome(self.path, chrome_options=self.chrome_options)
            cond=self.login()
            if cond==True:
                self.changepassprocess()
            else:
                return
        
        self.driver.get(str("https://m.trendyol.com/SifreGuncelleme"))
        
        oldpasselement = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div[2]/form/div/div[1]/div/div/input')))

        oldpasselement.send_keys(self.passw)
        newpasselement=self.driver.find_element_by_xpath('/html/body/main/div/div[2]/form/div/div[2]/div/div/input') #newpass
        newpasselement.send_keys(self.newpassw)
        newpasselement2=self.driver.find_element_by_xpath('/html/body/main/div/div[2]/form/div/div[3]/div/div/input') #newpass
        newpasselement2.send_keys(self.newpassw)
        button= WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div/div[2]/form/button')))#changebutton

        button.click()
        button= WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div[2]/div/button')))
        button.click()

        
    def managingpassporcess(self,username,newpass,oldpass):
        self.driver = webdriver.Chrome(self.path, chrome_options=self.chrome_options)
        self.username=username
        self.newpasslist=newpass
        self.oldpasslist=oldpass
        for i in range(0,len(self.userlist)):
            self.username=self.userlist[i]
            if self.pyqt.teksifre_chngoldpss.isChecked():
                self.passw=self.oldpasslist[0]
            else:
                self.passw=self.oldpasslist[i]
            if self.pyqt.teksifre_chngnewpss.isChecked():
                self.newpassw=self.newpasslist[0]
            else:
                self.newpassw=self.newpasslist[i]
            self.driver.close()
            self.driver = webdriver.Chrome(self.path, chrome_options=self.chrome_options)
            self.cond=self.login()
            
            if self.cond is True:
                self.changepassprocess()
            self.chngpss_step+=1
            self.pyqt.couponprogressBar_chngpss.setValue(self.chngpss_step)



















if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
    
     
